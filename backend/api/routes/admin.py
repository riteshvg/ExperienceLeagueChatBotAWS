"""
Admin endpoints — /api/admin/*

Authentication: POST /api/admin/login returns a short-lived JWT.
All other admin endpoints require that token as a Bearer header.
"""

import os
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Annotated, Optional

from fastapi import APIRouter, Depends, HTTPException, Request, status
from jose import jwt
from pydantic import BaseModel

_ROOT = Path(__file__).parent.parent.parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from backend.api.deps import get_admin_user, get_retriever, _secret, _ALGORITHM
from backend.core import google_db as _google_db
from config.settings import get_settings

router = APIRouter(prefix="/admin")

_TOKEN_TTL_HOURS = 1
_FEEDBACK_FILE = _ROOT / "data" / "feedback.jsonl"

# Citation stats are no longer tracked (citation_mapper removed); return empty.
_citation_stats: dict = {}


# ── Admin auth ────────────────────────────────────────────────────────────────

class LoginRequest(BaseModel):
    password: str


@router.post("/login")
async def login(body: LoginRequest):
    settings = get_settings()
    admin_password = settings.admin_password or ""
    if not admin_password or body.password != admin_password:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid password")
    # Use ADMIN_EMAIL as sub if set, so get_admin_user can verify it
    sub = (settings.admin_email or "").strip() or "admin"
    exp = datetime.now(tz=timezone.utc) + timedelta(hours=_TOKEN_TTL_HOURS)
    token = jwt.encode({"sub": sub, "exp": exp}, _secret(), algorithm=_ALGORITHM)
    return {"token": token, "expires_at": exp.isoformat()}


@router.post("/logout")
async def logout(_: Annotated[str, Depends(get_admin_user)]):
    return {"logged_out": True}


# ── Google OAuth user management ─────────────────────────────────────────────


class UpdateGoogleUserRequest(BaseModel):
    is_admin: Optional[bool] = None
    is_disabled: Optional[bool] = None


def _serialize_user(u: dict) -> dict:
    """Convert PostgreSQL row to JSON-safe dict (timestamps → ISO strings)."""
    result = {}
    for k, v in u.items():
        if hasattr(v, "isoformat"):
            result[k] = v.isoformat()
        elif isinstance(v, bool):
            result[k] = v
        else:
            result[k] = v
    return result


@router.get("/users")
async def list_users(_: Annotated[str, Depends(get_admin_user)]):
    """Return all Google OAuth users ordered by last_seen desc."""
    try:
        users = _google_db.list_users()
        return [_serialize_user(u) for u in users]
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"User DB unavailable: {exc}")


@router.patch("/users/{user_id}")
async def update_user(
    user_id: str,
    body: UpdateGoogleUserRequest,
    _: Annotated[str, Depends(get_admin_user)],
):
    """Update is_admin and/or is_disabled flags on a Google user."""
    try:
        updated: Optional[dict] = None
        if body.is_admin is not None:
            updated = _google_db.set_admin(user_id, body.is_admin)
        if body.is_disabled is not None:
            updated = _google_db.set_disabled(user_id, body.is_disabled)
        if updated is None:
            users = _google_db.list_users()
            updated = next((u for u in users if u["user_id"] == user_id), None)
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"User DB unavailable: {exc}")
    if not updated:
        raise HTTPException(status_code=404, detail="User not found")
    return _serialize_user(updated)


class UpdateLimitRequest(BaseModel):
    daily_query_limit: int


@router.patch("/users/{user_id}/limit")
async def update_user_limit(
    user_id: str,
    body: UpdateLimitRequest,
    _: Annotated[str, Depends(get_admin_user)],
):
    """Set per-user daily query limit."""
    updated = _google_db.set_user_daily_limit(user_id, body.daily_query_limit)
    if not updated:
        raise HTTPException(status_code=404, detail="User not found")
    return _serialize_user(updated)


@router.get("/users/summary")
async def users_summary(_: Annotated[str, Depends(get_admin_user)]):
    """Return aggregate stats: total users + total queries all time."""
    try:
        return _google_db.get_summary()
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"User DB unavailable: {exc}")


@router.get("/query-logs")
async def get_query_logs(
    _: Annotated[str, Depends(get_admin_user)],
    page: int = 1,
    page_size: int = 25,
    sort_by: str = "created_at",
    sort_order: str = "desc",
):
    """Return paginated query logs with sort and total count."""
    try:
        return _google_db.list_query_logs_paginated(
            page=max(1, page),
            page_size=min(max(1, page_size), 100),
            sort_by=sort_by,
            sort_order=sort_order,
        )
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Query logs unavailable: {exc}")


@router.get("/export/queries/excel")
async def export_queries_excel(
    _: Annotated[str, Depends(get_admin_user)],
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    """Export all query logs as an Excel file."""
    import io
    from datetime import date, datetime
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    try:
        records = _google_db.export_all_query_logs(date_from=date_from, date_to=date_to)
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Export failed: {exc}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Query Logs"

    headers = ["Time", "User Email", "Query", "Model", "Input Tokens", "Output Tokens", "Cost (USD)"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in records:
        created_at = r.get("created_at", "")
        try:
            dt = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
            created_at = dt.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            pass
        ws.append([
            created_at,
            r.get("email", ""),
            r.get("query_text", ""),
            r.get("llm_model", ""),
            r.get("input_tokens", 0),
            r.get("output_tokens", 0),
            round(float(r.get("cost_usd") or 0), 2),
        ])

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=row, column=col_idx).value or "")) for row in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 80)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"rovr-queries-{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Refresh pipeline ──────────────────────────────────────────────────────────

@router.get("/refresh/status")
async def refresh_status(_: Annotated[str, Depends(get_admin_user)]):
    from backend.core.refresh_pipeline import get_status
    return get_status()


@router.post("/refresh/start")
async def refresh_start(_: Annotated[str, Depends(get_admin_user)], force: bool = False):
    from backend.core.refresh_pipeline import trigger_refresh
    return trigger_refresh(force=force)


@router.post("/refresh/trigger-actions")
async def trigger_github_actions(_: Annotated[str, Depends(get_admin_user)], force: bool = False):
    import httpx
    token = os.getenv("GITHUB_TOKEN", "")
    if not token:
        raise HTTPException(status_code=503, detail="GITHUB_TOKEN not configured")
    url = "https://api.github.com/repos/riteshvg/ExperienceLeagueChatBotAWS/actions/workflows/refresh-docs.yml/dispatches"
    async with httpx.AsyncClient() as client:
        resp = await client.post(
            url,
            json={"ref": "main", "inputs": {"force": "true" if force else "false"}},
            headers={"Authorization": f"Bearer {token}", "Accept": "application/vnd.github.v3+json"},
            timeout=15,
        )
    if resp.status_code == 204:
        return {"triggered": True, "message": "GitHub Actions workflow dispatched"}
    raise HTTPException(status_code=resp.status_code, detail=f"GitHub API error: {resp.text}")


# ── Feedback ──────────────────────────────────────────────────────────────────

@router.get("/feedback")
async def get_feedback(_: Annotated[str, Depends(get_admin_user)]):
    try:
        entries = _google_db.list_feedback(limit=200)
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Feedback DB unavailable: {exc}")
    total = len(entries)
    thumbs_up = sum(1 for e in entries if e.get("rating") == 1)
    thumbs_down = sum(1 for e in entries if e.get("rating") == -1)
    return {
        "summary": {
            "total": total,
            "thumbs_up": thumbs_up,
            "thumbs_down": thumbs_down,
            "positive_pct": round(thumbs_up / total * 100, 1) if total else 0,
        },
        "entries": entries[:50],
    }


# ── Kill switch ───────────────────────────────────────────────────────────────

class KillSwitchRequest(BaseModel):
    enabled: bool


@router.get("/kill-switch")
async def get_kill_switch_status(_: Annotated[str, Depends(get_admin_user)]):
    from backend.core import kill_switch
    try:
        enabled = kill_switch.is_api_enabled()
    except Exception:
        enabled = True
    return {"enabled": enabled}


@router.post("/kill-switch")
async def set_kill_switch(body: KillSwitchRequest, _: Annotated[str, Depends(get_admin_user)]):
    from backend.core import kill_switch
    try:
        kill_switch.set_enabled(body.enabled)
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Kill switch unavailable: {exc}")
    return {"enabled": body.enabled}


# ── Demo ──────────────────────────────────────────────────────────────────────

@router.post("/demo/reset")
async def reset_demo(_: Annotated[str, Depends(get_admin_user)]):
    from backend.core.demo_counter import reset as demo_reset
    return demo_reset()


@router.get("/demo/status")
async def demo_status_admin(_: Annotated[str, Depends(get_admin_user)]):
    from backend.core.demo_counter import get_status
    return get_status()


# ── System status ─────────────────────────────────────────────────────────────

@router.get("/status")
async def system_status(request: Request, _: Annotated[str, Depends(get_admin_user)]):
    retriever = get_retriever(request)
    chroma_stats = retriever.collection_stats()
    product_breakdown = retriever.product_breakdown()

    bedrock_ok = True
    try:
        import boto3
        boto3.client("bedrock-runtime", region_name=os.getenv("BEDROCK_REGION", "us-east-1"))
    except Exception:
        bedrock_ok = False

    from backend.core.refresh_pipeline import get_status as get_refresh_status
    rs = get_refresh_status()

    return {
        "components": {
            "chromadb": {"healthy": True, **chroma_stats},
            "bedrock": {"healthy": bedrock_ok},
            "session_store": {
                "healthy": True,
                "active_sessions": len(request.app.state.session_store.list_sessions()),
            },
        },
        "citation_stats": dict(_citation_stats),
        "environment": os.getenv("ENVIRONMENT", "development"),
        "knowledge_base": {
            "last_refreshed": rs.get("last_run"),
            "total_chunks": chroma_stats.get("document_count", 0),
            "product_breakdown": product_breakdown,
        },
    }


@router.get("/settings")
async def get_settings_view(_: Annotated[str, Depends(get_admin_user)]):
    return {
        "bedrock_model_id": os.getenv("BEDROCK_MODEL_ID", "us.anthropic.claude-sonnet-4-20250514-v1:0"),
        "bedrock_region": os.getenv("BEDROCK_REGION", "us-east-1"),
        "similarity_threshold": float(os.getenv("SIMILARITY_THRESHOLD", "0.6")),
        "max_retrieval_results": int(os.getenv("MAX_RETRIEVAL_RESULTS", "8")),
        "min_retrieval_results": int(os.getenv("MIN_RETRIEVAL_RESULTS", "3")),
        "query_enhancement_enabled": os.getenv("QUERY_ENHANCEMENT_ENABLED", "true").lower() == "true",
        "environment": os.getenv("ENVIRONMENT", "development"),
    }


class DefaultLimitRequest(BaseModel):
    default_daily_limit: int


@router.get("/settings/default-limit")
async def get_default_limit(_: Annotated[str, Depends(get_admin_user)]):
    """Return the current default daily query limit from system_config."""
    raw = _google_db.get_system_config("default_daily_limit")
    return {"default_daily_limit": int(raw) if raw and raw.isdigit() else 20}


@router.patch("/settings/default-limit")
async def update_default_limit(body: DefaultLimitRequest, _: Annotated[str, Depends(get_admin_user)]):
    """Update the default daily limit for new users."""
    _google_db.set_system_config("default_daily_limit", str(body.default_daily_limit))
    return {"default_daily_limit": body.default_daily_limit}


@router.post("/settings/apply-default-limit")
async def apply_default_limit(_: Annotated[str, Depends(get_admin_user)]):
    """Apply the default_daily_limit to all existing users."""
    count = _google_db.apply_default_limit_to_all()
    raw = _google_db.get_system_config("default_daily_limit")
    return {"users_updated": count, "applied_limit": int(raw) if raw and raw.isdigit() else 20}


@router.get("/analytics")
async def analytics(request: Request, _: Annotated[str, Depends(get_admin_user)]):
    sessions = request.app.state.session_store.list_sessions()
    total_turns = sum(
        len(request.app.state.session_store.get_history(s)) // 2
        for s in sessions
    )
    rate_limits: dict = {}
    try:
        rate_limits = _google_db.get_rate_limit_analytics()
    except Exception:
        pass
    return {
        "active_sessions": len(sessions),
        "total_turns": total_turns,
        "citation_registry_hits": _citation_stats.get("registry_hits", 0),
        "citation_fallback_misses": _citation_stats.get("fallback_misses", 0),
        "rate_limits": rate_limits,
    }
